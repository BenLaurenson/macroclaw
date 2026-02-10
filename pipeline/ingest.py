"""Data ingestion module for MacroFactor .xlsx exports.

Handles reading Excel files exported from MacroFactor, detecting their type
based on column headers, mapping columns to the DuckDB schema, deduplication
via file hashing, and archival of processed files.
"""

import hashlib
import json
import logging
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd

from pipeline.schema import init_db

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column-header signatures used to auto-detect the export type.
# MacroFactor exports have recognisable header patterns.
# ---------------------------------------------------------------------------

_NUTRITION_SIGNATURES = {"Calories", "Protein", "Carbs", "Fat"}
_WORKOUT_SIGNATURES = {"Exercise Name", "Reps", "Weight"}
_WEIGHT_SIGNATURES = {"Scale Weight", "Trend Weight"}
_SUMMARY_SIGNATURES = {"Calorie Target", "Expenditure"}

# ---------------------------------------------------------------------------
# Sheets in the MacroFactor all-time (bulk) export that we can ingest.
# ---------------------------------------------------------------------------

_BULK_SHEET_MAP: dict[str, str] = {
    "Calories & Macros": "summary",
    "Scale Weight": "weight_scale",
    "Weight Trend": "weight_trend",
    "Expenditure": "expenditure",
    "Nutrition Program Settings": "targets",
}

# ---------------------------------------------------------------------------
# Column mappings: MacroFactor header -> DuckDB column
# ---------------------------------------------------------------------------

_NUTRITION_COL_MAP: dict[str, str] = {
    "Date": "date",
    "Meal": "meal",
    "Calories": "calories",
    "Protein": "protein_g",
    "Carbs": "carbs_g",
    "Fat": "fat_g",
    "Fiber": "fiber_g",
    "Sodium": "sodium_mg",
    "Food Name": "food_name",
    "Name": "food_name",
}

_WORKOUT_COL_MAP: dict[str, str] = {
    "Date": "date",
    "Workout Name": "workout_name",
    "Duration": "duration_min",
    "Exercise Name": "exercise_name",
    "Set Number": "set_number",
    "Set": "set_number",
    "Reps": "reps",
    "Weight": "weight_kg",
    "RPE": "rpe",
    "Notes": "notes",
}

_WEIGHT_COL_MAP: dict[str, str] = {
    "Date": "date",
    "Scale Weight": "scale_weight_kg",
    "Trend Weight": "trend_weight_kg",
}

_SUMMARY_COL_MAP: dict[str, str] = {
    "Date": "date",
    "Calories": "total_calories",
    "Total Calories": "total_calories",
    "Protein": "total_protein_g",
    "Total Protein": "total_protein_g",
    "Carbs": "total_carbs_g",
    "Total Carbs": "total_carbs_g",
    "Fat": "total_fat_g",
    "Total Fat": "total_fat_g",
    "Calorie Target": "calorie_target",
    "Protein Target": "protein_target_g",
    "Expenditure": "expenditure_kcal",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _file_hash(path: str) -> str:
    """Return the SHA-256 hex digest of a file's contents."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip unit suffixes like ``(kcal)``, ``(g)``, ``(kg)`` from column names.

    MacroFactor bulk (all-time) exports use headers like ``Calories (kcal)``
    and ``Protein (g)`` while daily exports use plain ``Calories`` and
    ``Protein``.  This normalises them to the plain form so the existing
    column maps work unchanged.
    """
    rename = {}
    for col in df.columns:
        cleaned = re.sub(r"\s*\([^)]*\)\s*$", "", col).strip()
        if cleaned != col:
            rename[col] = cleaned
    if rename:
        logger.debug("Normalised columns: %s", rename)
        df = df.rename(columns=rename)
    return df


def _detect_export_type(columns: list[str]) -> str:
    """Determine the MacroFactor export type from column headers.

    Args:
        columns: List of column names from the Excel file.

    Returns:
        One of ``"nutrition"``, ``"workout"``, ``"weight"``, or ``"summary"``.

    Raises:
        ValueError: If the column pattern does not match any known export type.
    """
    col_set = set(columns)

    # Order matters: summary and nutrition both have Calories/Protein, so check
    # summary-specific columns first.
    if _SUMMARY_SIGNATURES & col_set:
        return "summary"
    if _WORKOUT_SIGNATURES & col_set:
        return "workout"
    if _WEIGHT_SIGNATURES & col_set:
        return "weight"
    if _NUTRITION_SIGNATURES & col_set:
        return "nutrition"

    raise ValueError(
        f"Cannot detect export type from columns: {columns}. "
        "Expected MacroFactor nutrition, workout, weight, or summary export."
    )


def _map_columns(df: pd.DataFrame, col_map: dict[str, str]) -> pd.DataFrame:
    """Rename DataFrame columns using the provided mapping, dropping unmapped columns."""
    rename = {src: dst for src, dst in col_map.items() if src in df.columns}
    df = df.rename(columns=rename)
    # Keep only mapped columns that exist
    target_cols = list(rename.values())
    return df[[c for c in target_cols if c in df.columns]]


def _collect_extra_fields(
    row: pd.Series, known_cols: set[str]
) -> str | None:
    """Serialize unmapped columns into a JSON string for the food_details field."""
    extras = {k: v for k, v in row.items() if k not in known_cols and pd.notna(v)}
    return json.dumps(extras) if extras else None


def _prepare_nutrition(df: pd.DataFrame, source: str) -> pd.DataFrame:
    """Prepare a nutrition DataFrame for insertion."""
    mapped = _map_columns(df, _NUTRITION_COL_MAP)

    # Collect extra columns as JSON in food_details
    known = set(_NUTRITION_COL_MAP.values())
    mapped["food_details"] = df.apply(
        lambda row: _collect_extra_fields(row, set(_NUTRITION_COL_MAP.keys()) | known),
        axis=1,
    )

    mapped["source"] = source
    mapped["imported_at"] = datetime.now()

    # Ensure required columns exist even if empty
    for col in ["meal", "food_name"]:
        if col not in mapped.columns:
            mapped[col] = "Unknown"
    mapped["meal"] = mapped["meal"].fillna("Unknown")
    mapped["food_name"] = mapped["food_name"].fillna("Unknown")

    return mapped


def _prepare_workouts(df: pd.DataFrame, source: str) -> pd.DataFrame:
    """Prepare a workout DataFrame for insertion."""
    mapped = _map_columns(df, _WORKOUT_COL_MAP)
    mapped["source"] = source
    mapped["imported_at"] = datetime.now()

    if "set_number" not in mapped.columns:
        mapped["set_number"] = range(1, len(mapped) + 1)
    mapped["exercise_name"] = mapped["exercise_name"].fillna("Unknown")

    return mapped


def _prepare_weight(df: pd.DataFrame, source: str) -> pd.DataFrame:
    """Prepare a weight log DataFrame for insertion."""
    mapped = _map_columns(df, _WEIGHT_COL_MAP)
    mapped["source"] = source
    mapped["imported_at"] = datetime.now()
    return mapped


def _prepare_summary(df: pd.DataFrame, source: str) -> pd.DataFrame:
    """Prepare a daily summary DataFrame for insertion."""
    mapped = _map_columns(df, _SUMMARY_COL_MAP)
    mapped["source"] = source
    mapped["imported_at"] = datetime.now()
    return mapped


# ---------------------------------------------------------------------------
# Upsert logic
# ---------------------------------------------------------------------------

_UPSERT_SQL: dict[str, str] = {
    "nutrition": """
        INSERT OR REPLACE INTO nutrition_log
            (date, meal, calories, protein_g, carbs_g, fat_g, fiber_g,
             sodium_mg, food_name, food_details, source, imported_at)
        SELECT * FROM staging_df
    """,
    "workout": """
        INSERT OR REPLACE INTO workouts
            (date, workout_name, duration_min, exercise_name, set_number,
             reps, weight_kg, rpe, notes, source, imported_at)
        SELECT * FROM staging_df
    """,
    "weight": """
        INSERT OR REPLACE INTO weight_log
            (date, scale_weight_kg, trend_weight_kg, source, imported_at)
        SELECT * FROM staging_df
    """,
    "summary": """
        INSERT OR REPLACE INTO daily_summary
            (date, total_calories, total_protein_g, total_carbs_g, total_fat_g,
             calorie_target, protein_target_g, expenditure_kcal, source, imported_at)
        SELECT * FROM staging_df
    """,
}

# Expected column order for each table so SELECT * aligns correctly.
_EXPECTED_COLUMNS: dict[str, list[str]] = {
    "nutrition": [
        "date", "meal", "calories", "protein_g", "carbs_g", "fat_g",
        "fiber_g", "sodium_mg", "food_name", "food_details", "source", "imported_at",
    ],
    "workout": [
        "date", "workout_name", "duration_min", "exercise_name", "set_number",
        "reps", "weight_kg", "rpe", "notes", "source", "imported_at",
    ],
    "weight": [
        "date", "scale_weight_kg", "trend_weight_kg", "source", "imported_at",
    ],
    "summary": [
        "date", "total_calories", "total_protein_g", "total_carbs_g", "total_fat_g",
        "calorie_target", "protein_target_g", "expenditure_kcal", "source", "imported_at",
    ],
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def _is_bulk_export(xlsx_path: str) -> bool:
    """Return True if the workbook contains multiple MacroFactor sheets."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True)
    names = set(wb.sheetnames)
    wb.close()
    return bool(names & set(_BULK_SHEET_MAP))


def _upsert_df(
    conn: duckdb.DuckDBPyConnection,
    staging_df: pd.DataFrame,
    etype: str,
    source: str,
) -> int:
    """Upsert a prepared DataFrame into the appropriate table.

    Returns the number of rows inserted.
    """
    expected = _EXPECTED_COLUMNS[etype]
    for col in expected:
        if col not in staging_df.columns:
            staging_df[col] = None
    staging_df = staging_df[expected]

    staging_df["date"] = pd.to_datetime(staging_df["date"]).dt.date

    rows = len(staging_df)
    if rows == 0:
        return 0

    logger.info("Inserting %d rows into %s table", rows, etype)
    conn.register("staging_df", staging_df)
    conn.execute(_UPSERT_SQL[etype])
    conn.unregister("staging_df")
    return rows


def _apply_nutrition_targets(
    conn: duckdb.DuckDBPyConnection,
    targets_df: pd.DataFrame,
) -> int:
    """Resolve per-day calorie and protein targets from Nutrition Program Settings.

    MacroFactor stores targets as (program_update_date, weekday) pairs.
    For each date in daily_summary, we find the most recent program that was
    active on that date and look up the matching weekday row.

    Returns the number of daily_summary rows updated.
    """
    # Column names after normalisation (unit suffixes stripped)
    date_col = "Program Update Date"
    weekday_col = "Program Weekday"
    cal_col = "Calories"
    protein_col = "Protein"

    required = {date_col, weekday_col, cal_col, protein_col}
    if not required.issubset(targets_df.columns):
        logger.warning(
            "Nutrition Program Settings missing columns: %s",
            required - set(targets_df.columns),
        )
        return 0

    # Parse dates (DD/MM/YYYY format)
    targets_df = targets_df.copy()
    targets_df["_update_date"] = pd.to_datetime(
        targets_df[date_col], format="%d/%m/%Y"
    ).dt.date

    # Sort by update date descending so we can iterate newest-first
    update_dates = sorted(targets_df["_update_date"].unique(), reverse=True)

    # Build lookup: {(update_date, weekday_name) -> (calories, protein)}
    lookup: dict[tuple, tuple] = {}
    for _, row in targets_df.iterrows():
        key = (row["_update_date"], row[weekday_col])
        lookup[key] = (row[cal_col], row[protein_col])

    # Get all dates from daily_summary
    summary_dates = conn.execute(
        "SELECT date FROM daily_summary ORDER BY date"
    ).fetchall()

    weekday_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    updated = 0

    for (d,) in summary_dates:
        # Find the most recent program update <= this date
        active_update = None
        for ud in update_dates:
            if ud <= d:
                active_update = ud
                break

        if active_update is None:
            continue

        wday = weekday_names[d.weekday()]
        targets = lookup.get((active_update, wday))
        if targets is None:
            continue

        cal_target, protein_target = targets
        conn.execute(
            "UPDATE daily_summary SET calorie_target = ?, protein_target_g = ? WHERE date = ?",
            [cal_target, protein_target, d],
        )
        updated += 1

    logger.info("Updated %d daily_summary rows with calorie/protein targets", updated)
    return updated


def _ingest_bulk(
    conn: duckdb.DuckDBPyConnection,
    xlsx_path: str,
) -> dict[str, int]:
    """Process a MacroFactor all-time (bulk) export with multiple sheets.

    Returns a dict mapping table name to rows imported.
    """
    source = Path(xlsx_path).name
    stats: dict[str, int] = {}

    # --- Calories & Macros -> daily_summary (nutrition totals per day) ------
    try:
        macros_df = _normalize_columns(
            pd.read_excel(xlsx_path, sheet_name="Calories & Macros", engine="openpyxl")
        )
        if not macros_df.empty:
            summary_df = _prepare_summary(macros_df, source)
            stats["summary"] = _upsert_df(conn, summary_df, "summary", source)
    except Exception as e:
        logger.warning("Skipping 'Calories & Macros' sheet: %s", e)

    # --- Scale Weight + Weight Trend -> weight_log --------------------------
    try:
        scale_df = _normalize_columns(
            pd.read_excel(xlsx_path, sheet_name="Scale Weight", engine="openpyxl")
        )
        # Rename 'Weight' to 'Scale Weight' if needed (bulk uses "Weight (kg)")
        if "Weight" in scale_df.columns and "Scale Weight" not in scale_df.columns:
            scale_df = scale_df.rename(columns={"Weight": "Scale Weight"})
    except Exception as e:
        logger.warning("Skipping 'Scale Weight' sheet: %s", e)
        scale_df = pd.DataFrame()

    try:
        trend_df = _normalize_columns(
            pd.read_excel(xlsx_path, sheet_name="Weight Trend", engine="openpyxl")
        )
        if "Trend Weight" not in trend_df.columns:
            # Bulk export names it just "Trend Weight (kg)" -> normalized to "Trend Weight"
            for col in trend_df.columns:
                if col != "Date":
                    trend_df = trend_df.rename(columns={col: "Trend Weight"})
                    break
    except Exception as e:
        logger.warning("Skipping 'Weight Trend' sheet: %s", e)
        trend_df = pd.DataFrame()

    if not scale_df.empty or not trend_df.empty:
        # Merge scale and trend on Date
        if not scale_df.empty and not trend_df.empty:
            weight_df = pd.merge(scale_df[["Date", "Scale Weight"]], trend_df[["Date", "Trend Weight"]], on="Date", how="outer")
        elif not scale_df.empty:
            weight_df = scale_df[["Date", "Scale Weight"]].copy()
            weight_df["Trend Weight"] = None
        else:
            weight_df = trend_df[["Date", "Trend Weight"]].copy()
            weight_df["Scale Weight"] = None

        prepared = _prepare_weight(weight_df, source)
        stats["weight"] = _upsert_df(conn, prepared, "weight", source)

    # --- Expenditure -> merge into daily_summary ----------------------------
    try:
        exp_df = _normalize_columns(
            pd.read_excel(xlsx_path, sheet_name="Expenditure", engine="openpyxl")
        )
        if not exp_df.empty and "Expenditure" in exp_df.columns:
            exp_df["date"] = pd.to_datetime(exp_df["Date"]).dt.date
            for _, row in exp_df.iterrows():
                conn.execute(
                    "UPDATE daily_summary SET expenditure_kcal = ? WHERE date = ?",
                    [row["Expenditure"], row["date"]],
                )
            stats["expenditure_updates"] = len(exp_df)
    except Exception as e:
        logger.warning("Skipping 'Expenditure' sheet: %s", e)

    # --- Nutrition Program Settings -> calorie/protein targets ---------------
    try:
        targets_df = _normalize_columns(
            pd.read_excel(xlsx_path, sheet_name="Nutrition Program Settings", engine="openpyxl")
        )
        if not targets_df.empty:
            updates = _apply_nutrition_targets(conn, targets_df)
            stats["target_updates"] = updates
    except Exception as e:
        logger.warning("Skipping 'Nutrition Program Settings' sheet: %s", e)

    return stats


def ingest_xlsx(
    db_path: str,
    xlsx_path: str,
    export_type: str | None = None,
    archive_dir: str | None = None,
) -> dict[str, Any]:
    """Read a MacroFactor .xlsx export and load it into DuckDB.

    Automatically detects whether the file is a single-sheet daily export
    or a multi-sheet all-time (bulk) export.

    Args:
        db_path: Path to the DuckDB database file.
        xlsx_path: Path to the .xlsx file to ingest.
        export_type: One of ``"nutrition"``, ``"workout"``, ``"weight"``,
                     ``"summary"``, or ``None`` to auto-detect from headers.
        archive_dir: Directory to move processed files into.  When *None* the
                     file is left in place.

    Returns:
        A dict with keys ``export_type``, ``rows_imported``, ``file_hash``,
        ``file_path``, and ``skipped`` (bool).
    """
    xlsx_path = str(Path(xlsx_path).expanduser().resolve())
    logger.info("Starting ingestion of %s", xlsx_path)

    # Compute file hash for dedup
    fhash = _file_hash(xlsx_path)
    conn = init_db(db_path)

    # Check if already imported
    existing = conn.execute(
        "SELECT id FROM export_history WHERE file_hash = ?", [fhash]
    ).fetchone()
    if existing:
        logger.info("File already imported (hash=%s), skipping", fhash)
        conn.close()
        return {
            "export_type": export_type or "unknown",
            "rows_imported": 0,
            "file_hash": fhash,
            "file_path": xlsx_path,
            "skipped": True,
        }

    # Detect bulk vs single-sheet export
    if export_type is None and _is_bulk_export(xlsx_path):
        logger.info("Detected bulk (all-time) export — processing multiple sheets")
        sheet_stats = _ingest_bulk(conn, xlsx_path)
        total_rows = sum(sheet_stats.values())

        conn.execute(
            "INSERT INTO export_history (export_type, file_path, file_hash, rows_imported) "
            "VALUES (?, ?, ?, ?)",
            ["bulk", xlsx_path, fhash, total_rows],
        )
        conn.close()
        logger.info("Bulk ingestion complete: %s", sheet_stats)

        if archive_dir:
            _archive_file(xlsx_path, archive_dir)

        return {
            "export_type": "bulk",
            "rows_imported": total_rows,
            "file_hash": fhash,
            "file_path": xlsx_path,
            "skipped": False,
            "sheet_stats": sheet_stats,
        }

    # Single-sheet export path
    df = pd.read_excel(xlsx_path, engine="openpyxl")
    df = _normalize_columns(df)
    logger.info("Read %d rows and %d columns from %s", len(df), len(df.columns), xlsx_path)

    if df.empty:
        logger.warning("Empty spreadsheet: %s", xlsx_path)
        conn.close()
        return {
            "export_type": export_type or "unknown",
            "rows_imported": 0,
            "file_hash": fhash,
            "file_path": xlsx_path,
            "skipped": False,
        }

    # Detect or validate export type
    detected = _detect_export_type(df.columns.tolist())
    if export_type is not None and export_type != detected:
        logger.warning(
            "Specified export_type=%s but detected=%s; using specified value",
            export_type,
            detected,
        )
    etype = export_type or detected

    # Prepare data
    preparers = {
        "nutrition": _prepare_nutrition,
        "workout": _prepare_workouts,
        "weight": _prepare_weight,
        "summary": _prepare_summary,
    }
    source = Path(xlsx_path).name
    staging_df = preparers[etype](df, source)

    rows_imported = _upsert_df(conn, staging_df, etype, source)

    # Record in export history
    conn.execute(
        "INSERT INTO export_history (export_type, file_path, file_hash, rows_imported) "
        "VALUES (?, ?, ?, ?)",
        [etype, xlsx_path, fhash, rows_imported],
    )

    conn.close()
    logger.info("Ingestion complete: %d rows into %s", rows_imported, etype)

    if archive_dir:
        _archive_file(xlsx_path, archive_dir)

    return {
        "export_type": etype,
        "rows_imported": rows_imported,
        "file_hash": fhash,
        "file_path": xlsx_path,
        "skipped": False,
    }


def _archive_file(xlsx_path: str, archive_dir: str) -> None:
    """Move a processed file to the archive directory."""
    archive_dir = str(Path(archive_dir).expanduser().resolve())
    Path(archive_dir).mkdir(parents=True, exist_ok=True)
    dest = Path(archive_dir) / Path(xlsx_path).name
    if dest.exists():
        stem = dest.stem
        suffix = dest.suffix
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = dest.with_name(f"{stem}_{ts}{suffix}")
    shutil.move(xlsx_path, str(dest))
    logger.info("Archived file to %s", dest)
